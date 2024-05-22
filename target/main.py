import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
import red_envelope
import logging

# Configure logging
logging.basicConfig(filename='red_envelope_simulation.log', level=logging.ERROR,
                    format='%(asctime)s %(levelname)s:%(message)s')

# Constants
simulation_times = 1000000
max_people = 50
envelope_money = 20000  # This represents 200 yuan
excel_filename = 'red_envelope_simulation.xlsx'

# Create a workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Simulation Results'

# Set the header
ws['A1'] = f'Share {envelope_money / 100:.2f} yuan | simulate {simulation_times} times'
ws.append([])  # Empty line
ws.append(['Total Person number'] + [f'Person {i + 1}' for i in range(max_people)])

# Perform the simulation for each number of people from 1 to max_people
for n in range(1, max_people + 1):
    expected_distribution = red_envelope.analyze_red_envelope(simulation_times, n)
    row = [n] + [
        f'{amt / 100:.2f}' if i < n else 'x'
        for i, amt in enumerate(expected_distribution + [0] * (max_people - n))
    ]
    ws.append(row)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except (TypeError, ValueError) as e:
                logging.error(f"Error processing cell value: {cell.value}", exc_info=True)
    adjusted_width = max_length + 2
    ws.column_dimensions[column].width = adjusted_width

# Center align all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center')

# Save the workbook
wb.save(excel_filename)

# Create bar charts for 5, 10, 20, and 50 people
fig, axs = plt.subplots(2, 2, figsize=(15, 10))
people_cases = [5, 10, 20, 50]
colors = ['#009999', '#66CC00', '#FF8000', '#FF3333']

for idx, people in enumerate(people_cases):
    row = ws.iter_rows(min_row=4 + people - 1, max_row=4 + people - 1, values_only=True)
    amounts = next(row)[1:people + 1]
    ax = axs[idx // 2, idx % 2]
    ax.bar(range(1, people + 1), [float(amt) for amt in amounts if amt != 'x'], color=colors[idx])
    ax.set_title(f'{people} People Distribution')
    ax.set_xlabel('Person')
    ax.set_ylabel('Expectation Amount (yuan)')
    ax.set_xticks(range(1, people + 1))  # Ensure x-axis shows integers

# Adjust layout and save the figure
plt.tight_layout()
plt.savefig('distribution_barcharts.png')
plt.show()
